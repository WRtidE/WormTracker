"""WormTracker 主窗口

PyQt6 界面，提供视频选择、参数调节、实时预览、Excel 导出等完整功能。
"""

import os
from typing import Optional

import cv2
import numpy as np
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QLCDNumber, QTableWidget, QTableWidgetItem, QHeaderView, QSizePolicy,
    QPushButton, QFileDialog, QMessageBox, QGroupBox, QStatusBar, QFormLayout,
    QSpinBox, QDoubleSpinBox, QTabWidget,
)
from PyQt6.QtGui import QImage, QPixmap, QColor
from PyQt6.QtCore import Qt

from wormtracker.config import WormTrackerConfig, get_config, load_config, save_config
from wormtracker.engine.base import BaseEngine
from wormtracker.engine.mog2 import MOG2Engine
from wormtracker.ui.thread import VideoThread
from wormtracker.ui.styles import APP_STYLESHEET, APP_STYLESHEET_LIGHT


class WormCounterApp(QMainWindow):
    """WormTracker 主窗口"""

    # —— UI 反转：spinbox 显示"距顶部高度百分比"(越大越靠上)，内部转 ratio -—
    @staticmethod
    def _to_display(ratio: float) -> float:
        return round(1.0 - ratio, 4)

    @staticmethod
    def _from_display(display: float) -> float:
        return round(1.0 - display, 4)

    def __init__(self, config: Optional[WormTrackerConfig] = None):
        super().__init__()
        self.config = config or get_config()
        self.setWindowTitle("WormTracker — 自动化线虫计数系统 v1.0")
        self.setMinimumSize(750, 550)
        self.setStyleSheet(APP_STYLESHEET_LIGHT)
        self._is_dark = False

        # 运行时状态
        self.thread: Optional[VideoThread] = None
        self.engine: Optional[BaseEngine] = None
        self.current_counts: dict = {}
        self.current_video_path: str = ""
        self.labels: list[int] = sorted(
            range(1, self.config.num_channels + 1), reverse=True
        )
        self.preview_frame: Optional[np.ndarray] = None
        self.last_main_frame: Optional[np.ndarray] = None
        self.last_mask_frame: Optional[np.ndarray] = None
        self._show_recognition_view = False

        self._build_ui()
        self._reset_ui()

    # ==========================================================
    # UI 构建
    # ==========================================================

    def _build_ui(self) -> None:
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        # ---- 左侧: 视频预览 / 结果图表 ----
        self.video_container = QVBoxLayout()
        self.video_label = QLabel("等待导入实验视频...\n(请从右侧面板选择视频文件)")
        self.video_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.video_label.setStyleSheet(
            "background-color: #E8E3DA; color: #968B81; font-size: 20px; "
            "font-weight: bold; border: 2px dashed #CCC3B5; border-radius: 12px;"
        )
        self.video_label.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Ignored
        )
        self.video_container.addWidget(self.video_label)
        main_layout.addLayout(self.video_container, stretch=7)

        # ---- 右侧: 标签页 ----
        self.tabs = QTabWidget()
        self.tabs.addTab(self._build_tab_workbench(), "🎮 工作台")
        self.tabs.addTab(self._build_tab_params(), "⚙️ 参数调节")
        main_layout.addWidget(self.tabs, stretch=3)

        # 状态栏
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("✅ 系统就绪，请先选择实验视频。")

        # 主题切换按钮 (状态栏右侧)
        self.btn_theme = QPushButton("🌓")
        self.btn_theme.setToolTip("切换浅色/暗色主题")
        self.btn_theme.clicked.connect(self._toggle_theme)
        self.btn_theme.setStyleSheet(
            "QPushButton { background: transparent; border: none; font-size: 14px;"
            "  padding: 1px 4px; }"
            "QPushButton:hover { background-color: #D9D1C5; border-radius: 4px; }"
        )
        self.statusBar.addPermanentWidget(self.btn_theme)

        # 识别视角切换按钮
        self.btn_view = QPushButton("🔍")
        self.btn_view.setToolTip("切换识别视角 (查看模型检测到的前景/运动区域)")
        self.btn_view.clicked.connect(self._toggle_recognition_view)
        self.btn_view.setStyleSheet(
            "QPushButton { background: transparent; border: none; font-size: 14px;"
            "  padding: 1px 4px; }"
            "QPushButton:hover { background-color: #D9D1C5; border-radius: 4px; }"
        )
        self.statusBar.addPermanentWidget(self.btn_view)

    def _build_tab_workbench(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)

        # 系统控制
        ctrl_group = QGroupBox("🚀 系统控制")
        ctrl_layout = QVBoxLayout(ctrl_group)

        self.btn_open = self._create_btn("📂 选择实验视频", "#6B8BAE", "#FFFFFF")
        self.btn_open.clicked.connect(self._open_video)
        self.btn_main_action = self._create_btn(
            "▶ 请先导入视频", "#D9D1C5", "#968B81"
        )
        self.btn_main_action.clicked.connect(self._handle_main_action)
        self.btn_main_action.setEnabled(False)

        ctrl_layout.addWidget(self.btn_open)
        ctrl_layout.addWidget(self.btn_main_action)
        layout.addWidget(ctrl_group)

        # 实时数据
        data_group = QGroupBox("📊 实时数据")
        data_layout = QVBoxLayout(data_group)

        self.total_lcd = QLCDNumber()
        self.total_lcd.setDigitCount(5)
        self.total_lcd.setFixedHeight(60)
        self.total_lcd.setSegmentStyle(QLCDNumber.SegmentStyle.Flat)
        self.total_lcd.setStyleSheet(
            "color: #6B9F6E; background-color: #E8E3DA; "
            "border: 2px solid #D9D1C5; border-radius: 8px;"
        )

        self.table = QTableWidget(len(self.labels), 2)
        self.table.setHorizontalHeaderLabels(["通道 ID", "计数"])
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        self._label_total_header = QLabel(
            "<center style='color:#968B81; font-size:11px;'>TOTAL WORM COUNT</center>"
        )
        self._label_dist_header = QLabel(
            "<center style='color:#968B81; font-size:11px;'>CHANNEL DISTRIBUTION</center>"
        )
        data_layout.addWidget(self._label_total_header)
        data_layout.addWidget(self.total_lcd)
        data_layout.addWidget(self._label_dist_header)
        data_layout.addWidget(self.table)
        layout.addWidget(data_group, stretch=1)

        # 完成面板
        self.finish_group = QGroupBox("✅ 分析完成")
        self.finish_group.setStyleSheet(
            "QGroupBox { border-color: #6B9F6E; color: #6B9F6E; }"
        )
        finish_layout = QVBoxLayout(self.finish_group)
        self.btn_export = self._create_btn(
            "📊 导出 Excel 报告", "#6B9F6E", "#FFFFFF"
        )
        self.btn_export.clicked.connect(self._export_results)
        self.btn_export_save_config = self._create_btn(
            "💾 保存当前配置", "#6B8BAE", "#FFFFFF"
        )
        self.btn_export_save_config.clicked.connect(self._save_config_dialog)
        finish_layout.addWidget(self.btn_export)
        finish_layout.addWidget(self.btn_export_save_config)
        self.finish_group.hide()
        layout.addWidget(self.finish_group)

        return tab

    def _draw_chart(self, counts: dict) -> None:
        """在视频预览区展示结果柱状图"""

        # ── 延迟导入 matplotlib (首次显示图表时才加载，加快启动速度) ──
        import matplotlib
        matplotlib.use("QtAgg")
        import matplotlib.font_manager as fm
        import platform as _plat
        _sys = _plat.system()
        if _sys == "Darwin":
            _candidate_fonts = ['PingFang HK', 'Songti SC', 'STHeiti', 'Heiti TC', 'Apple LiGothic']
        elif _sys == "Windows":
            _candidate_fonts = ['Microsoft YaHei', 'SimHei', 'SimSun', 'KaiTi', 'FangSong']
        else:
            _candidate_fonts = ['WenQuanYi Micro Hei', 'Noto Sans CJK SC', 'DejaVu Sans']
        _available = {f.name for f in fm.fontManager.ttflist}
        _for_use = [f for f in _candidate_fonts if f in _available]
        if _for_use:
            matplotlib.rcParams['font.sans-serif'] = _for_use + ['DejaVu Sans']
            matplotlib.rcParams['axes.unicode_minus'] = False
            fm._load_fontmanager(try_read_cache=False)
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.figure import Figure

        sorted_labels = sorted(counts.keys())
        values = [counts[k] for k in sorted_labels]
        channel_names = [f"CH-{k:02d}" for k in sorted_labels]
        total = sum(values)

        max_val = max(values) if max(values) > 0 else 1

        # 主题自适应
        d = self._is_dark
        fig_bg   = "#1e1e2e" if d else "#F0ECE5"
        ax_bg    = "#11111b" if d else "#EDE8E0"
        edge_c   = "#313244" if d else "#D9D1C5"
        text_c   = "#cdd6f4" if d else "#5C534A"
        label_c  = "#a6adc8" if d else "#968B81"
        title_c  = "#a6e3a1" if d else "#6B9F6E"
        spine_c  = "#45475a" if d else "#CCC3B5"
        grid_c   = "#313244" if d else "#D9D1C5"

        fig = Figure(figsize=(8, 5), dpi=100)
        fig.patch.set_facecolor(fig_bg)
        ax = fig.add_subplot(111)
        ax.set_facecolor(ax_bg)

        # 渐变色柱状图 (绿→黄)
        colors = []
        for v in values:
            ratio = v / max_val
            r = int(100 + 155 * (1 - ratio))
            g = int(255 * ratio + 150 * (1 - ratio))
            b = int(100 * (1 - ratio))
            colors.append(f"#{r:02x}{g:02x}{b:02x}")

        bars = ax.bar(range(len(values)), values, color=colors, edgecolor=edge_c, linewidth=0.5)

        # 数值标签
        for bar, val in zip(bars, values):
            if val > 0:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + max_val * 0.03,
                    str(val), ha="center", va="bottom",
                    fontsize=8, color=text_c, fontweight="bold",
                )

        ax.set_xticks(range(len(values)))
        xticks_rotation = 45 if len(values) > 16 else 0
        ax.set_xticklabels(channel_names, rotation=xticks_rotation,
                           ha="right" if xticks_rotation else "center", fontsize=7)
        ax.set_ylabel("线虫数量", color=label_c, fontsize=9)
        ax.set_title(f"各通道线虫计数  |  总计: {total} 条  |  峰值通道: CH-{sorted_labels[values.index(max_val)]:02d}",
                     color=title_c, fontsize=12, fontweight="bold", pad=10)
        ax.tick_params(colors=label_c, labelsize=7)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["bottom"].set_color(spine_c)
        ax.spines["left"].set_color(spine_c)
        ax.yaxis.grid(True, color=grid_c, linewidth=0.5, alpha=0.7)
        ax.set_axisbelow(True)

        fig.tight_layout(pad=1.5)

        # 嵌入到视频区域
        canvas = FigureCanvas(fig)
        canvas.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.video_label.hide()
        self.video_container.addWidget(canvas)
        self._result_canvas = canvas

    def _restore_video_area(self) -> None:
        """恢复视频预览 (重新分析时调用)"""
        if hasattr(self, "_result_canvas") and self._result_canvas is not None:
            self._result_canvas.hide()
            self.video_container.removeWidget(self._result_canvas)
            self._result_canvas.setParent(None)
            self._result_canvas = None
        self.video_label.show()

    def _build_tab_params(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        param_group = QGroupBox("⚙️ 核心参数 (实时生效)")
        param_layout = QFormLayout(param_group)
        param_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # 通道数量
        self.spin_channels = QSpinBox()
        self.spin_channels.setFixedHeight(40)
        self.spin_channels.setRange(1, 100)
        self.spin_channels.setValue(self.config.num_channels)
        self.spin_channels.valueChanged.connect(self._on_channels_changed)
        param_layout.addRow("微流控通道数量:", self.spin_channels)

        # 检测线高度 (显示为距顶部百分比，越大越靠上)
        self.spin_tripwire = QDoubleSpinBox()
        self.spin_tripwire.setFixedHeight(40)
        self.spin_tripwire.setRange(0.1, 0.9)
        self.spin_tripwire.setSingleStep(0.02)
        self.spin_tripwire.setValue(self._to_display(self.config.tripwire_ratio))
        self.spin_tripwire.valueChanged.connect(self._on_tripwire_changed)
        param_layout.addRow("检测线高度 (红线):", self.spin_tripwire)

        # Mask 上界 (显示为距顶部百分比，越大越靠上)
        self.spin_mask_top = QDoubleSpinBox()
        self.spin_mask_top.setFixedHeight(40)
        self.spin_mask_top.setRange(0.0, 1.0)
        self.spin_mask_top.setSingleStep(0.02)
        self.spin_mask_top.setValue(self._to_display(self.config.mask_top_ratio))
        self.spin_mask_top.valueChanged.connect(self._on_mask_changed)
        param_layout.addRow("Mask 上界 (屏蔽上方):", self.spin_mask_top)

        # Mask 下界 (显示为距顶部百分比，越大越靠上)
        self.spin_mask_bottom = QDoubleSpinBox()
        self.spin_mask_bottom.setFixedHeight(40)
        self.spin_mask_bottom.setRange(0.0, 1.0)
        self.spin_mask_bottom.setSingleStep(0.02)
        self.spin_mask_bottom.setValue(self._to_display(self.config.mask_bottom_ratio))
        self.spin_mask_bottom.valueChanged.connect(self._on_mask_changed)
        param_layout.addRow("Mask 下界 (屏蔽下方):", self.spin_mask_bottom)

        # 最小虫体面积
        self.spin_min_area = QSpinBox()
        self.spin_min_area.setFixedHeight(40)
        self.spin_min_area.setRange(5, 500)
        self.spin_min_area.setSingleStep(10)
        self.spin_min_area.setValue(self.config.min_area)
        self.spin_min_area.valueChanged.connect(
            lambda v: self._update_thread_param("min_area", v)
        )
        param_layout.addRow("虫体滤噪 (最小面积):", self.spin_min_area)

        layout.addWidget(param_group)

        self._tips_label = QLabel(
            "💡 提示：在【暂停】或【未开始】时，\n"
            "修改高度可直接在画面预览检测线位置。\n"
            "更改通道数量将实时重置表格。"
        )
        self._tips_label.setStyleSheet(
            "color: #6B8BAE; font-weight: normal; line-height: 1.5;"
        )
        self._tips_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self._tips_label)
        layout.addStretch(1)

        return tab

    # ==========================================================
    # 交互逻辑
    # ==========================================================

    def _create_btn(
        self, text: str, bg_color: Optional[str] = None, text_color: Optional[str] = None
    ) -> QPushButton:
        btn = QPushButton(text)
        if bg_color and text_color:
            btn.setStyleSheet(
                f"QPushButton {{ background-color: {bg_color}; color: {text_color}; }} "
                f"QPushButton:hover {{ filter: brightness(110%); border: 1px solid white; }} "
                f"QPushButton:disabled {{ {self._disabled_style()} }}"
            )
        return btn

    # ── 主题切换 ──────────────────────────────────────

    def _disabled_style(self) -> str:
        if self._is_dark:
            return "background-color: #181825; color: #585b70; border: 1px solid #313244;"
        return "background-color: #EDE8E0; color: #B8AFA6; border: 1px solid #D9D1C5;"

    def _toggle_theme(self) -> None:
        self._apply_theme(not self._is_dark)

    def _toggle_recognition_view(self) -> None:
        """切换识别视角"""
        self._show_recognition_view = not self._show_recognition_view
        if self._show_recognition_view:
            self.btn_view.setText("👁️")
            self.statusBar.showMessage("🔍 识别视角：显示模型检测到的前景/运动区域")
        else:
            self.btn_view.setText("🔍")
            self.statusBar.showMessage("✅ 系统就绪，请先选择实验视频。")
        # 恢复最近一帧
        if self.last_main_frame is not None:
            self._show_latest_frame()

    def _apply_theme(self, is_dark: bool) -> None:
        self._is_dark = is_dark
        self.setStyleSheet(APP_STYLESHEET if is_dark else APP_STYLESHEET_LIGHT)
        self._refresh_theme_widgets()

    def _refresh_theme_widgets(self) -> None:
        """刷新所有内联样式控件"""
        d = self._is_dark

        # 视频占位区
        bg = "#11111b" if d else "#E8E3DA"
        fg = "#585b70" if d else "#968B81"
        bd = "#45475a" if d else "#CCC3B5"
        self.video_label.setStyleSheet(
            f"background-color: {bg}; color: {fg}; font-size: 20px; "
            f"font-weight: bold; border: 2px dashed {bd}; border-radius: 12px;"
        )

        # LCD 总数
        lcd_color = "#a6e3a1" if d else "#6B9F6E"
        lcd_bg = "#11111b" if d else "#E8E3DA"
        lcd_border = "#313244" if d else "#D9D1C5"
        self.total_lcd.setStyleSheet(
            f"color: {lcd_color}; background-color: {lcd_bg}; "
            f"border: 2px solid {lcd_border}; border-radius: 8px;"
        )

        # 完成面板边框
        finish_color = "#a6e3a1" if d else "#6B9F6E"
        self.finish_group.setStyleSheet(
            f"QGroupBox {{ border-color: {finish_color}; color: {finish_color}; }}"
        )

        # 提示文字 (参数页)
        if hasattr(self, "_tips_label"):
            tip_color = "#89b4fa" if d else "#6B8BAE"
            self._tips_label.setStyleSheet(
                f"color: {tip_color}; font-weight: normal; line-height: 1.5;"
            )

        # 列头标签
        header_color = "#a6adc8" if d else "#968B81"
        if hasattr(self, "_label_total_header"):
            self._label_total_header.setText(
                f"<center style='color:{header_color}; font-size:11px;'>TOTAL WORM COUNT</center>"
            )
        if hasattr(self, "_label_dist_header"):
            self._label_dist_header.setText(
                f"<center style='color:{header_color}; font-size:11px;'>CHANNEL DISTRIBUTION</center>"
            )

        # 静态功能按钮
        blue_bg  = "#89b4fa" if d else "#6B8BAE"
        blue_txt = "#11111b" if d else "#FFFFFF"
        green_bg  = "#a6e3a1" if d else "#6B9F6E"
        green_txt = "#11111b" if d else "#FFFFFF"
        gray_bg  = "#313244" if d else "#D9D1C5"
        gray_txt = "#a6adc8" if d else "#968B81"

        self.btn_open.setStyleSheet(
            f"QPushButton {{ background-color: {blue_bg}; color: {blue_txt}; }} "
            f"QPushButton:hover {{ filter: brightness(110%); border: 1px solid white; }} "
            f"QPushButton:disabled {{ {self._disabled_style()} }}"
        )
        self.btn_export.setStyleSheet(
            f"QPushButton {{ background-color: {green_bg}; color: {green_txt}; }} "
            f"QPushButton:hover {{ filter: brightness(110%); border: 1px solid white; }} "
            f"QPushButton:disabled {{ {self._disabled_style()} }}"
        )
        self.btn_export_save_config.setStyleSheet(
            f"QPushButton {{ background-color: {blue_bg}; color: {blue_txt}; }} "
            f"QPushButton:hover {{ filter: brightness(110%); border: 1px solid white; }} "
            f"QPushButton:disabled {{ {self._disabled_style()} }}"
        )

        if not self.btn_main_action.isEnabled():
            self.btn_main_action.setStyleSheet(
                f"QPushButton {{ background-color: {gray_bg}; color: {gray_txt}; }} "
                f"QPushButton:disabled {{ {self._disabled_style()} }}"
            )

        # 主操作按钮 (根据当前文字恢复对应样式)
        self._refresh_main_action_btn()

        # 主题按钮 hover
        hover_bg = "#45475a" if d else "#D9D1C5"
        for btn in [self.btn_theme, self.btn_view]:
            btn.setStyleSheet(
                "QPushButton { background: transparent; border: none; font-size: 14px;"
                "  padding: 1px 4px; }"
                f"QPushButton:hover {{ background-color: {hover_bg}; border-radius: 4px; }}"
            )

        # 图表 (如果正在显示)
        if hasattr(self, "_result_canvas") and self._result_canvas is not None:
            # 重建图表
            if self.current_counts:
                self._draw_chart(self.current_counts)

    def _refresh_main_action_btn(self) -> None:
        """根据按钮文字恢复主题匹配的内联样式"""
        d = self._is_dark
        text = self.btn_main_action.text()
        if "继续" in text:
            color = "#a6e3a1" if d else "#6B9F6E"
            text_c = "#11111b" if d else "#FFFFFF"
        elif "暂停" in text:
            color = "#f9e2af" if d else "#C4844E"
            text_c = "#11111b" if d else "#FFFFFF"
        elif "重新分析" in text:
            color = "#89b4fa" if d else "#6B8BAE"
            text_c = "#11111b" if d else "#FFFFFF"
        elif "开始识别" in text:
            color = "#a6e3a1" if d else "#6B9F6E"
            text_c = "#11111b" if d else "#FFFFFF"
        else:
            return
        self.btn_main_action.setStyleSheet(
            f"background-color: {color}; color: {text_c};"
        )

    def _update_thread_param(self, key: str, value) -> None:
        if self.thread and self.thread.isRunning():
            self.thread.update_param(key, value)

    def _on_channels_changed(self, value: int) -> None:
        self._update_thread_param("num_channels", value)
        self.labels = sorted(range(1, value + 1), reverse=True)
        new_counts = {label: self.current_counts.get(label, 0) for label in self.labels}
        self.current_counts = new_counts

        self.table.setRowCount(len(self.labels))
        for i, label in enumerate(self.labels):
            item_id = QTableWidgetItem(f"CH-{label:02d}")
            item_id.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 0, item_id)
            item_count = QTableWidgetItem(str(self.current_counts[label]))
            item_count.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 1, item_count)

    def _on_tripwire_changed(self, value: float) -> None:
        ratio = self._from_display(value)
        self._update_thread_param("tripwire_ratio", ratio)
        if (
            self.thread is None
            or not self.thread.isRunning()
            or self.thread.is_paused
        ):
            self._render_preview_frame()

    def _on_mask_changed(self) -> None:
        self._update_thread_param("mask_top_ratio", self._from_display(self.spin_mask_top.value()))
        self._update_thread_param("mask_bottom_ratio", self._from_display(self.spin_mask_bottom.value()))
        if (
            self.thread is None
            or not self.thread.isRunning()
            or self.thread.is_paused
        ):
            self._render_preview_frame()

    def _render_preview_frame(self) -> None:
        img_to_draw = None
        if self.thread and self.thread.isRunning() and self.thread.is_paused:
            if self.last_main_frame is not None:
                img_to_draw = self.last_main_frame.copy()
        elif not (self.thread and self.thread.isRunning()):
            if self.preview_frame is not None:
                img_to_draw = self.preview_frame.copy()

        if img_to_draw is not None:
            h, w = img_to_draw.shape[:2]
            # display 值转 ratio 再算 Y 像素
            trip_ratio = self._from_display(self.spin_tripwire.value())
            trip_y = int(h * trip_ratio)
            cv2.line(img_to_draw, (0, trip_y), (w, trip_y), (0, 0, 255), 3)
            cv2.putText(
                img_to_draw,
                f"Tripwire: {self.spin_tripwire.value():.2f}",
                (15, trip_y - 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2,
            )

            mt_y = int(h * self._from_display(self.spin_mask_top.value()))
            mb_y = int(h * self._from_display(self.spin_mask_bottom.value()))
            cv2.line(img_to_draw, (0, mt_y), (w, mt_y), (0, 215, 255), 2)
            cv2.putText(
                img_to_draw,
                f"Mask Top: {self.spin_mask_top.value():.2f}",
                (15, mt_y - 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 215, 255), 1,
            )
            cv2.line(img_to_draw, (0, mb_y), (w, mb_y), (0, 215, 255), 2)
            cv2.putText(
                img_to_draw,
                f"Mask Bot: {self.spin_mask_bottom.value():.2f}",
                (15, mb_y - 10),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 215, 255), 1,
            )

            self._display_frame(img_to_draw)

    def _display_frame(self, frame: np.ndarray) -> None:
        h, w = frame.shape[:2]
        fmt = (
            QImage.Format.Format_RGB888
            if len(frame.shape) == 3
            else QImage.Format.Format_Grayscale8
        )
        qt_img = QImage(frame.data, w, h, frame.strides[0], fmt)
        if len(frame.shape) == 3:
            qt_img = qt_img.rgbSwapped()
        pixmap = QPixmap.fromImage(qt_img).scaled(
            self.video_label.width(), self.video_label.height(),
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        self.video_label.setPixmap(pixmap)

    def keyPressEvent(self, event) -> None:
        if event.key() == Qt.Key.Key_Space and self.btn_main_action.isEnabled():
            self._handle_main_action()
        super().keyPressEvent(event)

    # ==========================================================
    # 动作处理
    # ==========================================================

    def _handle_main_action(self) -> None:
        if self.thread is None or not self.thread.isRunning():
            self._start_analysis()
        elif self.thread.isRunning() and not self.thread.is_paused:
            self.thread.toggle_pause()
            self.btn_main_action.setText("▶ 继续分析 (Space)")
            self._refresh_main_action_btn()
            self.statusBar.showMessage("⏸ 分析已暂停，可切换到【参数调节】预览检测线。")
        elif self.thread.isRunning() and self.thread.is_paused:
            self.thread.toggle_pause()
            self.btn_main_action.setText("⏸ 暂停分析 (Space)")
            self._refresh_main_action_btn()
            self.statusBar.showMessage("▶️ 正在实时分析中...")
        self.setFocus()

    def _open_video(self) -> None:
        if self.thread and self.thread.isRunning():
            reply = QMessageBox.question(
                self, "中止确认",
                "当前正在分析视频，是否要强制中止并导入新视频？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.No:
                return
            self.thread.stop()
            self.thread = None

        file_name, _ = QFileDialog.getOpenFileName(
            self, "选择实验视频", "",
            "视频 (*.mp4 *.avi *.mov *.mkv)"
        )
        if file_name:
            self.current_video_path = file_name
            self._reset_ui()
            self.finish_group.hide()

            self.btn_main_action.setEnabled(True)
            self.btn_main_action.setText("▶ 开始识别")
            self._refresh_main_action_btn()
            self.statusBar.showMessage(
                f"📂 已加载: {os.path.basename(file_name)}。"
                "你可以切换到【参数调节】预览检测线高度。"
            )

            cap = cv2.VideoCapture(file_name)
            ret, frame = cap.read()
            if ret:
                self.preview_frame = frame.copy()
                self._render_preview_frame()
            else:
                self.video_label.setText(
                    "成功导入，但预览画面失败\n请直接点击【开始识别】"
                )
            cap.release()

    def _start_analysis(self) -> None:
        if not self.current_video_path:
            return
        self._restore_video_area()
        self._reset_ui()
        self.finish_group.hide()

        self.btn_main_action.setText("⏸ 暂停分析 (Space)")
        self._refresh_main_action_btn()
        # 创建 MOG2 引擎
        self.engine = MOG2Engine(self.config)

        # 创建线程
        self.thread = VideoThread(
            self.current_video_path, self.engine, self.config
        )
        self.thread.update_param("num_channels", self.spin_channels.value())
        self.thread.update_param("tripwire_ratio", self._from_display(self.spin_tripwire.value()))
        self.thread.update_param("min_area", self.spin_min_area.value())
        self.thread.update_param("mask_top_ratio", self._from_display(self.spin_mask_top.value()))
        self.thread.update_param("mask_bottom_ratio", self._from_display(self.spin_mask_bottom.value()))

        self.thread.change_pixmap_signal.connect(self._update_frame)
        self.thread.update_counts_signal.connect(self._update_stats)
        self.thread.finished_signal.connect(self._on_finished)
        self.thread.start()

        self.tabs.setCurrentIndex(0)
        self.statusBar.showMessage("▶️ 正在实时分析中...")
        self.setFocus()

    def _on_finished(self) -> None:
        self.btn_main_action.setEnabled(True)
        self.btn_main_action.setText("🔄 重新分析该视频")
        self._refresh_main_action_btn()
        self.finish_group.show()

        # 在视频区域展示结果图表
        if self.current_counts:
            self._draw_chart(self.current_counts)
        self.tabs.setCurrentIndex(0)
        self.statusBar.showMessage(
            "✅ 视频分析完毕，图表已展示在左侧。可导出报告或重新分析。"
        )

    def _export_results(self) -> None:
        if not self.current_counts:
            return
        base_name = os.path.splitext(os.path.basename(self.current_video_path))[0]
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel 报告",
            f"{base_name}",
            "Excel Files (*.xlsx)",
        )
        if save_path:
            try:
                sorted_keys = sorted(self.current_counts.keys())
                video_name = os.path.basename(self.current_video_path)
                gen_time = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")

                # 主数据表
                data = {
                    "微通道编号": [f"CH-{k:02d}" for k in sorted_keys],
                    "线虫统计数量": [self.current_counts[k] for k in sorted_keys],
                }
                df = pd.DataFrame(data)
                total_row = pd.DataFrame({
                    "微通道编号": ["总计 (Total)"],
                    "线虫统计数量": [sum(self.current_counts.values())],
                })
                df = pd.concat([df, total_row], ignore_index=True)

                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    # 将视频名称和生成时间作为表头信息写入"计数结果"表
                    import openpyxl
                    from openpyxl.utils import get_column_letter

                    def _auto_width(ws, min_width=10, max_width=50):
                        """根据内容自适应列宽，中文字符按 2 倍宽度计算"""
                        for col_cells in ws.columns:
                            max_len = 0
                            col_letter = get_column_letter(col_cells[0].column)
                            for cell in col_cells:
                                val = str(cell.value) if cell.value is not None else ""
                                # CJK 字符视为双倍宽度
                                char_len = sum(2 if '\u4e00' <= c <= '\u9fff' or '\u3000' <= c <= '\u303f' or '\uff00' <= c <= '\uffef' else 1 for c in val)
                                max_len = max(max_len, char_len)
                            adjusted = max(min_width, min(max_len + 3, max_width))
                            ws.column_dimensions[col_letter].width = adjusted

                    df.to_excel(writer, index=False, sheet_name="计数结果", startrow=3)
                    ws = writer.sheets["计数结果"]
                    ws["A1"] = "视频名称"
                    ws["B1"] = video_name
                    ws["A2"] = "生成时间"
                    ws["B2"] = gen_time
                    # 设置表头信息样式
                    ws["A1"].font = openpyxl.styles.Font(bold=True, color="333333")
                    ws["A2"].font = openpyxl.styles.Font(bold=True, color="333333")
                    _auto_width(ws)

                    # 实验信息表
                    pd.DataFrame({
                        "项目": ["视频名称", "生成时间", "检测后端"],
                        "内容": [
                            video_name,
                            gen_time,
                            self.config.backend.upper(),
                        ],
                    }).to_excel(writer, index=False, sheet_name="实验信息")
                    _auto_width(writer.sheets["实验信息"])

                self.statusBar.showMessage(f"💾 报告已成功导出至: {save_path}")
                QMessageBox.information(self, "导出成功", "Excel 报表已生成！")
            except Exception as e:
                QMessageBox.critical(self, "导出失败", str(e))

    def _save_config_dialog(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "保存配置文件", "wormtracker_config.yaml",
            "YAML (*.yaml *.yml)",
        )
        if path:
            save_config(path, self.config)
            self.statusBar.showMessage(f"💾 配置已保存至: {path}")

    # ==========================================================
    # UI 更新
    # ==========================================================

    def _reset_ui(self) -> None:
        self.total_lcd.display(0)
        self.current_counts = {label: 0 for label in self.labels}
        self.table.setRowCount(len(self.labels))
        for i, label in enumerate(self.labels):
            item_id = QTableWidgetItem(f"CH-{label:02d}")
            item_id.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 0, item_id)
            item_count = QTableWidgetItem("0")
            item_count.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 1, item_count)

    def _update_frame(
        self, main_frame: np.ndarray, mask_frame: np.ndarray
    ) -> None:
        self.last_main_frame = main_frame.copy()
        self.last_mask_frame = mask_frame.copy()
        self._show_latest_frame()

    def _show_latest_frame(self) -> None:
        if self._show_recognition_view and self.last_mask_frame is not None:
            self._display_frame(self._build_recognition_frame(self.last_mask_frame))
        elif self.last_main_frame is not None:
            self._display_frame(self.last_main_frame)

    @staticmethod
    def _build_recognition_frame(mask: np.ndarray) -> np.ndarray:
        """将二值前景 mask 渲染为彩色识别视图

        热点区 (白/灰) → 亮青绿色
        静止区 (纯黑) → 暗蓝背景
        """
        norm = mask.astype(np.float32) / 255.0
        # 热点 -> 荧光绿黄; 背景 -> 深蓝灰
        h, w = mask.shape[:2]
        out = np.zeros((h, w, 3), dtype=np.uint8)
        out[:, :, 0] = (norm * 20).astype(np.uint8)         # B: 暗背景偏蓝
        out[:, :, 1] = (norm * 240).astype(np.uint8)        # G: 热点亮绿
        out[:, :, 2] = (norm * 160 + (1 - norm) * 15).astype(np.uint8)  # R: 热点偏黄
        return out

    def _update_stats(self, counts: dict) -> None:
        self.current_counts = counts
        self.total_lcd.display(sum(counts.values()))
        for i, label in enumerate(self.labels):
            val = str(counts.get(label, 0))
            item = self.table.item(i, 1)
            if item and item.text() != val:
                item.setText(val)
                item.setForeground(QColor("#a6e3a1" if self._is_dark else "#6B9F6E"))

    def closeEvent(self, event) -> None:
        if self.thread:
            self.thread.stop()
        event.accept()
